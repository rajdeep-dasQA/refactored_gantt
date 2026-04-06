"""
Input reader: parses the input .xlsx file and returns a Project instance.

Expected sheet layout
---------------------
Row  | Column A         | Column B  | ...
-----|------------------|-----------|----
 1   | Start Date       | 2025-01-06
 2   | Hours Per Day    | 6
 ...
 N   | Summary          | Assignee  | Priority | Estimated Hours   ← header sentinel
 N+1 | <task row>       | ...
 ... | ...
     | TOTAL HOURS      | ...       ← ignored / stops reading
"""
from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

from gantt.models.project import Project
from gantt.models.task import Task


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def read_project(path: str | Path) -> Project:
    """Parse *path* and return a fully-configured (unscheduled) Project."""
    wb = load_workbook(str(path))
    ws = wb.active

    config   = _parse_config(ws)
    tasks    = _parse_tasks(ws)

    return Project(
        start_date    = config["start_date"],
        hours_per_day = config["hours_per_day"],
        tasks         = tasks,
    )


# ---------------------------------------------------------------------------
# Private helpers
# ---------------------------------------------------------------------------

_SKIP_SUMMARIES = {"TOTAL", "TOTAL HOURS", ""}
_DATE_FMT       = "%Y-%m-%d"


def _parse_config(ws) -> dict:
    config: dict = {"start_date": None, "hours_per_day": 6.0}

    for row in ws.iter_rows(values_only=True):
        if not row or row[0] is None:
            continue
        key = str(row[0]).strip().lower()

        if key == "start date":
            val = row[1]
            config["start_date"] = (
                val if isinstance(val, datetime)
                else datetime.strptime(str(val).strip(), _DATE_FMT)
            )
        elif key == "hours per day":
            config["hours_per_day"] = float(row[1])
        elif key == "summary":
            break   # reached the task-table header

    if config["start_date"] is None:
        raise ValueError("Input sheet must contain a 'Start Date' config row.")

    return config


def _parse_tasks(ws) -> list[Task]:
    tasks:   list[Task] = []
    reading: bool       = False
    order:   int        = 0

    for row in ws.iter_rows(values_only=True):
        if not row or row[0] is None:
            continue
        key = str(row[0]).strip().lower()

        if key == "summary":
            reading = True
            continue

        if not reading:
            continue

        summary = str(row[0]).strip()
        if summary.upper() in _SKIP_SUMMARIES:
            continue

        assignee = str(row[1]).strip() if row[1] else ""

        try:
            priority = int(row[2]) if row[2] is not None else 99
        except (ValueError, TypeError):
            priority = 99

        try:
            hours = float(row[3]) if row[3] is not None else 0.0
        except (ValueError, TypeError):
            continue   # skip rows that cannot be parsed as tasks

        tasks.append(Task(
            summary  = summary,
            assignee = assignee,
            priority = priority,
            hours    = hours,
            _order   = order,
        ))
        order += 1

    # Stable sort: priority first, then original row order within same priority
    tasks.sort(key=lambda t: (t.priority, t._order))
    return tasks
