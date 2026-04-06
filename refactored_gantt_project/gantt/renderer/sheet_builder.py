"""
Builds a single Gantt worksheet inside an existing openpyxl Workbook.
Called once for the daily view and once for the weekly view.
"""
from __future__ import annotations
from typing import Literal

from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

from gantt.config.constants import (
    ALT_ROW, BAR_COLORS, BORDER_C, COL_W_ASSIGNEE, COL_W_DAY_PERIOD,
    COL_W_HOURS, COL_W_PRIORITY, COL_W_SUMMARY, COL_W_WEEK_PERIOD,
    DEFAULT_BAR, DEFAULT_COLOR, HDR_BG, HDR_FG, PRIORITY_COLORS,
    ROW_H_HEADER, ROW_H_LEGEND, ROW_H_TASK, ROW_H_TITLE, ROW_H_TOTAL, WHITE,
)
from gantt.models.project import Project
from gantt.renderer.styles import (
    apply_header_style, make_border, make_fill, make_font,
    make_thick_left_border,
)
from gantt.scheduler.calendar_utils import Period, get_periods
from gantt.scheduler.engine import hours_in_period

ViewMode = Literal["day", "week"]


def build_sheet(
    wb:      Workbook,
    project: Project,
    mode:    ViewMode,
) -> object:   # openpyxl Worksheet
    """
    Create and populate one Gantt worksheet in *wb*.

    Returns the created worksheet.
    """
    periods = get_periods(project.start_date, project.project_end, mode)

    icon  = "📆" if mode == "week" else "📅"
    label = "Weekly" if mode == "week" else "Daily"
    sheet_title = f"{icon}  {label} View"

    ws = wb.create_sheet(sheet_title)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "E3"

    _set_column_widths(ws, periods, mode)
    _write_title_row(ws, project, icon, label, periods)
    _write_header_row(ws, periods)
    _write_task_rows(ws, project, periods)
    _write_total_row(ws, project, periods)
    _write_legend(ws, project, periods)

    return ws


# ---------------------------------------------------------------------------
# Private helpers
# ---------------------------------------------------------------------------

def _set_column_widths(ws, periods: list[Period], mode: ViewMode) -> None:
    ws.column_dimensions["A"].width = COL_W_SUMMARY
    ws.column_dimensions["B"].width = COL_W_ASSIGNEE
    ws.column_dimensions["C"].width = COL_W_PRIORITY
    ws.column_dimensions["D"].width = COL_W_HOURS
    col_w = COL_W_WEEK_PERIOD if mode == "week" else COL_W_DAY_PERIOD
    for i in range(len(periods)):
        ws.column_dimensions[get_column_letter(5 + i)].width = col_w


def _write_title_row(ws, project: Project, icon: str, label: str, periods: list[Period]) -> None:
    ws.row_dimensions[1].height = ROW_H_TITLE
    last_col = get_column_letter(4 + len(periods))
    ws.merge_cells(f"A1:{last_col}1")
    c = ws["A1"]
    c.value = (
        f"{icon}  {label} View   |   "
        f"Start: {project.start_date.strftime('%d %b %Y')}   |   "
        f"{project.hours_per_day}h/day   |   "
        f"Total: {project.total_hours:.0f}h"
    )
    c.font      = make_font(bold=True, color=HDR_FG, size=11)
    c.fill      = make_fill(HDR_BG)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)


def _write_header_row(ws, periods: list[Period]) -> None:
    ws.row_dimensions[2].height = ROW_H_HEADER
    static_headers = [
        ("Summary",                  "center"),
        ("Assignee",                 "center"),
        ("Priority",                 "center"),
        ("Original\nEstimate\n(hrs)","center"),
    ]
    for ci, (text, align) in enumerate(static_headers, start=1):
        c = ws.cell(row=2, column=ci, value=text)
        apply_header_style(c, bg=HDR_BG, h_align=align, wrap=True)

    for i, (label, _, _) in enumerate(periods):
        c = ws.cell(row=2, column=5 + i, value=label)
        apply_header_style(c, bg=HDR_BG, size=8, h_align="center", wrap=True)


def _write_task_rows(ws, project: Project, periods: list[Period]) -> None:
    for ri, task in enumerate(project.tasks):
        row    = 3 + ri
        ws.row_dimensions[row].height = ROW_H_TASK
        alt    = (ri % 2 == 1)
        row_bg = ALT_ROW if alt else WHITE

        pri                  = task.priority
        bar_color            = BAR_COLORS.get(pri, DEFAULT_BAR)
        label_bg, label_fg   = PRIORITY_COLORS.get(pri, DEFAULT_COLOR)

        # --- Summary (coloured by priority) ----------------------------------
        c = ws.cell(row=row, column=1, value=task.summary)
        c.font      = make_font(bold=True, color=label_fg, size=9)
        c.fill      = make_fill(label_bg)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        c.border    = make_thick_left_border()

        # --- Assignee --------------------------------------------------------
        c = ws.cell(row=row, column=2, value=task.assignee)
        c.font      = make_font(size=9)
        c.fill      = make_fill(row_bg)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        c.border    = make_border()

        # --- Priority --------------------------------------------------------
        c = ws.cell(row=row, column=3, value=task.priority)
        c.font      = make_font(size=9)
        c.fill      = make_fill(row_bg)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = make_border()

        # --- Estimated hours -------------------------------------------------
        c = ws.cell(row=row, column=4, value=task.hours)
        c.font      = make_font(bold=True, size=9)
        c.fill      = make_fill(row_bg)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = make_border()

        # --- Period bar cells ------------------------------------------------
        for i, (_, p_start, p_end) in enumerate(periods):
            hrs = hours_in_period(task, p_start, p_end, project.hours_per_day)
            c   = ws.cell(row=row, column=5 + i)
            c.border = make_border()
            if hrs > 0:
                display = int(hrs) if hrs == int(hrs) else hrs
                c.value     = display
                c.font      = make_font(bold=True, color="FFFFFF", size=9)
                c.fill      = make_fill(bar_color)
                c.alignment = Alignment(horizontal="center", vertical="center")
            else:
                c.value = ""
                c.fill  = make_fill(row_bg)


def _write_total_row(ws, project: Project, periods: list[Period]) -> None:
    total_row = 3 + len(project.tasks)
    ws.row_dimensions[total_row].height = ROW_H_TOTAL

    ws.merge_cells(f"A{total_row}:C{total_row}")
    c = ws.cell(row=total_row, column=1, value="TOTAL HOURS")
    c.font      = make_font(bold=True, color=HDR_FG, size=10)
    c.fill      = make_fill(HDR_BG)
    c.alignment = Alignment(horizontal="right", vertical="center", indent=1)
    c.border    = make_border()

    c2 = ws.cell(row=total_row, column=4, value=f"=SUM(D3:D{total_row - 1})")
    c2.font      = make_font(bold=True, color=HDR_FG, size=10)
    c2.fill      = make_fill(HDR_BG)
    c2.alignment = Alignment(horizontal="center", vertical="center")
    c2.border    = make_border()

    for i in range(len(periods)):
        c = ws.cell(row=total_row, column=5 + i)
        c.fill   = make_fill(HDR_BG)
        c.border = make_border()


def _write_legend(ws, project: Project, periods: list[Period]) -> None:
    total_row = 3 + len(project.tasks)
    leg1 = total_row + 1
    leg2 = total_row + 2

    ws.row_dimensions[leg1].height = ROW_H_LEGEND
    ws.merge_cells(f"A{leg1}:D{leg1}")
    c = ws[f"A{leg1}"]
    c.value     = "Cell value = hours worked in that period   |   Color = priority level"
    c.font      = make_font(italic=True, size=8, color="888888")
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    ws.row_dimensions[leg2].height = ROW_H_LEGEND
    for i, (pri, (bg, fg)) in enumerate(PRIORITY_COLORS.items()):
        c = ws.cell(row=leg2, column=1 + i, value=f"Priority {pri}")
        c.font      = make_font(bold=True, color=fg, size=8)
        c.fill      = make_fill(bg)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = make_border()
