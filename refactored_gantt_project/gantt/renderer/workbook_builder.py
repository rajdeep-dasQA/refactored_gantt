"""
Workbook-level orchestration.

* generate_gantt_workbook  – create the output .xlsx from a Project
* generate_input_template  – write a blank (sample-filled) input .xlsx
"""
from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Alignment

from gantt.config.constants import (
    ALT_ROW, BORDER_C, HDR_BG, HDR_FG, INFO_BD, INFO_BG, INFO_FG,
    ROW_H_CONFIG, WHITE,
)
from gantt.models.project import Project
from gantt.renderer.sheet_builder import build_sheet
from gantt.renderer.styles import (
    make_border, make_fill, make_font,
)

# ---------------------------------------------------------------------------
# Gantt output workbook
# ---------------------------------------------------------------------------

def generate_gantt_workbook(project: Project, output_path: str) -> None:
    """
    Render both Daily and Weekly view sheets into a single workbook
    and save to *output_path*.
    """
    wb = Workbook()
    wb.remove(wb.active)          # drop the default empty sheet

    day_ws  = build_sheet(wb, project, mode="day")
    _       = build_sheet(wb, project, mode="week")   # noqa: F841

    wb.active = day_ws            # open on Daily view by default
    wb.save(output_path)


# ---------------------------------------------------------------------------
# Input template workbook
# ---------------------------------------------------------------------------

SAMPLE_TASKS = [
    ("Project Setup and Configuration",                "Dibyawan Trivedi", 1, 6),
    ("Project Setup & Configuration",                  "Somnath Bera",     1, 10),
    ("DB Schema, Repository and Migrations",           "Somnath Bera",     1, 20),
    ("New Badge variants",                             "Dibyawan Trivedi", 2, 2),
    ("New Stepper Form Generic Component",             "Dibyawan Trivedi", 2, 4),
    ("New Tab Design",                                 "Dibyawan Trivedi", 2, 2),
    ("Add New Button Variant",                         "Dibyawan Trivedi", 2, 2),
    ("Implement Contact Stage UI & Validation",        "Dibyawan Trivedi", 3, 8),
    ("Implement Target Stage UI & Validation",         "Dibyawan Trivedi", 3, 6),
    ("Page Navigation & Data Persistence",             "Dibyawan Trivedi", 3, 4),
    ("Implement Target Stage Endpoint and Validation", "Somnath Bera",     3, 8),
    ("Contact Stage API and validation Implementation","Somnath Bera",     3, 5),
    ("Layout Initialization",                          "Dibyawan Trivedi", 5, 4),
    ("KPI Dashboard Integration",                      "Dibyawan Trivedi", 5, 4),
    ("Leads Data Table",                               "Dibyawan Trivedi", 5, 8),
    ("Navigation & Routing",                           "Dibyawan Trivedi", 5, 4),
]


def generate_input_template(output_path: str) -> None:
    """Write a sample-filled input template to *output_path*."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Input"
    ws.sheet_view.showGridLines = False

    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 28

    # ── Title ──────────────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 30
    ws.merge_cells("A1:E1")
    c = ws["A1"]
    c.value     = "📋  GANTT INPUT SHEET"
    c.font      = make_font(bold=True, color=HDR_FG, size=13)
    c.fill      = make_fill(HDR_BG)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    # ── Config rows ────────────────────────────────────────────────────────
    config_data = [
        ("Start Date",    "2025-01-06", "Format: YYYY-MM-DD"),
        ("Hours Per Day", 6,            "Productive hours per workday"),
    ]
    for ri, (key, val, note) in enumerate(config_data, start=2):
        ws.row_dimensions[ri].height = ROW_H_CONFIG
        _config_key_cell(ws, ri, 1, key)
        _config_val_cell(ws, ri, 2, val)
        _note_cell(ws, ri, 3, note)

    # ── Info banner ────────────────────────────────────────────────────────
    ws.row_dimensions[4].height = 22
    ws.merge_cells("A4:E4")
    c = ws["A4"]
    c.value     = "💡  Both Daily and Weekly views are auto-generated — click the sheet tabs to switch between them."
    c.font      = make_font(italic=True, color=INFO_FG, size=9)
    c.fill      = make_fill(INFO_BG)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    c.border    = make_border(INFO_BD)

    # spacer
    ws.row_dimensions[5].height = 6

    # ── Task table header ──────────────────────────────────────────────────
    ws.row_dimensions[6].height = 24
    for ci, h in enumerate(["Summary", "Assignee", "Priority", "Estimated Hours"], start=1):
        c = ws.cell(row=6, column=ci, value=h)
        c.font      = make_font(bold=True, color=HDR_FG, size=10)
        c.fill      = make_fill(HDR_BG)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = make_border()
    ws.cell(row=6, column=5).value = "← Row order within same priority = execution sequence"
    ws.cell(row=6, column=5).font  = make_font(italic=True, color="888888", size=9)

    # ── Sample task rows ───────────────────────────────────────────────────
    for ri, (summ, assignee, pri, hrs) in enumerate(SAMPLE_TASKS, start=7):
        ws.row_dimensions[ri].height = 18
        bg = ALT_ROW if ri % 2 == 0 else WHITE
        for ci, val in enumerate([summ, assignee, pri, hrs], start=1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.font      = make_font(size=9)
            c.fill      = make_fill(bg)
            c.alignment = Alignment(
                horizontal="left" if ci < 3 else "center",
                vertical="center",
                indent=1 if ci < 3 else 0,
            )
            c.border = make_border()

    # ── Total row ──────────────────────────────────────────────────────────
    total_row = 7 + len(SAMPLE_TASKS)
    ws.row_dimensions[total_row].height = 22
    ws.merge_cells(f"A{total_row}:C{total_row}")
    c = ws.cell(row=total_row, column=1, value="TOTAL HOURS")
    c.font      = make_font(bold=True, color=HDR_FG, size=10)
    c.fill      = make_fill(HDR_BG)
    c.alignment = Alignment(horizontal="right", vertical="center", indent=1)
    c.border    = make_border()

    c2 = ws.cell(row=total_row, column=4, value=f"=SUM(D7:D{total_row - 1})")
    c2.font      = make_font(bold=True, color=HDR_FG, size=11)
    c2.fill      = make_fill(HDR_BG)
    c2.alignment = Alignment(horizontal="center", vertical="center")
    c2.border    = make_border()

    wb.save(output_path)


# ---------------------------------------------------------------------------
# Cell helpers
# ---------------------------------------------------------------------------

def _config_key_cell(ws, row: int, col: int, value: str) -> None:
    c = ws.cell(row=row, column=col, value=value)
    c.font      = make_font(bold=True, color=HDR_FG, size=10)
    c.fill      = make_fill(HDR_BG)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    c.border    = make_border()


def _config_val_cell(ws, row: int, col: int, value) -> None:
    c = ws.cell(row=row, column=col, value=value)
    c.font      = make_font(bold=True, size=10)
    c.fill      = make_fill("EEF3F8")
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    c.border    = make_border()


def _note_cell(ws, row: int, col: int, note: str) -> None:
    c = ws.cell(row=row, column=col, value=note)
    c.font      = make_font(italic=True, color="888888", size=9)
    c.alignment = Alignment(horizontal="left", vertical="center")
