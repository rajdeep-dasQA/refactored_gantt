"""
openpyxl style-object factories.

All colour arguments are plain 6-char hex strings (no leading 'FF').
The helpers below add the alpha prefix internally.
"""
from __future__ import annotations

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from gantt.config.constants import BORDER_C, FONT_NAME


# ---------------------------------------------------------------------------
# Low-level helpers
# ---------------------------------------------------------------------------

def col_hex(h: str) -> str:
    """Convert 6-char hex to openpyxl ARGB format (prepend 'FF')."""
    return "FF" + h.upper().lstrip("#")


def make_fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=col_hex(hex_color))


def make_font(
    bold:   bool = False,
    color:  str  = "000000",
    size:   int  = 10,
    italic: bool = False,
) -> Font:
    return Font(
        bold=bold,
        color=col_hex(color),
        size=size,
        italic=italic,
        name=FONT_NAME,
    )


def make_border(color: str = BORDER_C) -> Border:
    side = Side(style="thin", color=col_hex(color))
    return Border(left=side, right=side, top=side, bottom=side)


def make_thick_left_border(color: str = BORDER_C) -> Border:
    """Medium left edge + thin on all other sides (used for the Summary column)."""
    thick = Side(style="medium", color=col_hex("999999"))
    thin  = Side(style="thin",   color=col_hex(color))
    return Border(left=thick, right=thin, top=thin, bottom=thin)


# ---------------------------------------------------------------------------
# Compound style helper
# ---------------------------------------------------------------------------

def apply_header_style(
    cell,
    bg:      str,
    fg:      str  = "FFFFFF",
    size:    int  = 10,
    bold:    bool = True,
    h_align: str  = "center",
    wrap:    bool = False,
) -> None:
    """Apply a standard header look to *cell* in-place."""
    cell.font      = make_font(bold=bold, color=fg, size=size)
    cell.fill      = make_fill(bg)
    cell.alignment = Alignment(
        horizontal=h_align, vertical="center", wrap_text=wrap, indent=(1 if h_align == "left" else 0)
    )
    cell.border    = make_border()
